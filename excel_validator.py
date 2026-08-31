import csv
import json
import os
import re
from pathlib import Path

import requests
from openpyxl.styles import PatternFill

EXCEL_PATH = os.getenv("APARTMENT_EXCEL_PATH", "Central_NJ_Apartment_Comparison_WITH_TOUR_CLUSTERS_FINAL.xlsx")
MAX_PRICE = int(os.getenv("MAX_PRICE", "3200"))
COMMUTE_LIMIT = int(os.getenv("COMMUTE_LIMIT", "30"))
REQUIRED_BEDS = int(os.getenv("MIN_BEDS", "2"))
REQUIRED_BATHS = int(os.getenv("MIN_BATHS", "2"))
TAVILY_API_KEY = os.getenv("TAVILY_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")


def normalize_field(value):
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize_header(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


def get_field_value(row, *names):
    lookup = {normalize_header(k): v for k, v in row.items()}
    for name in names:
        if normalize_header(name) in lookup:
            return lookup[normalize_header(name)]
    return None


def parse_number(value):
    if value is None:
        return None
    match = re.search(r"\d[\d,]*(?:\.\d+)?", str(value))
    if not match:
        return None
    return float(match.group(0).replace(",", ""))


def parse_commute_minutes(value):
    text = normalize_field(value)
    if not text:
        return None
    match = re.search(r"(\d{1,2})(?:\s*[-–]\s*(\d{1,2}))?\s*(?:min|mins|minute|minutes)", text, re.IGNORECASE)
    if not match:
        return None
    lower = int(match.group(1))
    upper = match.group(2)
    if upper:
        return (lower + int(upper)) / 2
    return float(lower)


def row_is_managed_community(row):
    text = normalize_field(
        " ".join(
            [
                get_field_value(row, "Management") or "",
                get_field_value(row, "Property Type") or "",
                get_field_value(row, "Community Name", "Name") or "",
                get_field_value(row, "Notes") or "",
            ]
        )
    ).lower()
    if not text:
        return False
    managed_markers = [
        "managed",
        "leasing office",
        "property management",
        "apartment community",
        "townhome community",
    ]
    return any(marker in text for marker in managed_markers)


def repair_row(row):
    repaired = {}
    for key, value in row.items():
        cleaned = normalize_field(value)
        if normalize_header(key) in {"community name", "property", "name", "community"}:
            repaired[key] = cleaned
        elif normalize_header(key) in {"price", "monthly rent", "estimated 2 bed price range"}:
            repaired[key] = cleaned if cleaned else "Not listed"
        elif normalize_header(key) in {"beds", "bedroom", "bedrooms"}:
            repaired[key] = cleaned if cleaned else "Not listed"
        elif normalize_header(key) in {"baths", "bathroom", "bathrooms"}:
            repaired[key] = cleaned if cleaned else "Not listed"
        elif normalize_header(key) in {"property type", "type"}:
            repaired[key] = cleaned if cleaned else "Apartment"
        elif normalize_header(key) in {"address", "location", "community address"}:
            repaired[key] = cleaned if cleaned else "Not listed"
        elif normalize_header(key) in {"commute", "drive time", "commute time"}:
            repaired[key] = cleaned if cleaned else "Not listed"
        elif normalize_header(key) in {"leasing office phone", "phone", "contact"}:
            repaired[key] = cleaned if cleaned else "Not listed"
        elif normalize_header(key) in {"official website", "website", "link", "url"}:
            repaired[key] = cleaned if cleaned else "Not listed"
        else:
            repaired[key] = cleaned
    return repaired


def validate_row(row):
    issues = []
    normalized = {k: normalize_field(v) for k, v in row.items()} 

    community_name = (
        get_field_value(row, "Community Name", "Property", "Name")
        or ""
    )
    community_name = normalize_field(community_name)
    price_candidates = [
        parse_number(get_field_value(row, "Price", "Monthly Rent", "Estimated 2-Bed Price Range")),
    ]
    price_values = [n for n in price_candidates if n is not None]
    if not community_name:
        issues.append("Missing community name.")
    if not row_is_managed_community(row):
        issues.append("Community is not clearly identified as a managed property.")
    if not get_field_value(row, "Address", "Location"):
        issues.append("Missing address.")
    if not get_field_value(row, "Beds", "Bedroom"):
        issues.append("Missing bedroom count.")
    if not get_field_value(row, "Baths", "Bathroom"):
        issues.append("Missing bathroom count.")
    if not get_field_value(row, "Property Type", "Type"):
        issues.append("Missing property type.")
    if not get_field_value(row, "Amenities", "Key Amenities"):
        issues.append("Missing amenities.")
    if not get_field_value(row, "Leasing Office Phone", "Phone", "Contact"):
        issues.append("Missing leasing office phone/contact.")
    if not get_field_value(row, "Official Website", "Website", "Link", "URL"):
        issues.append("Missing official website or booking link.")

    bed_fields = [
        get_field_value(row, "Beds", "Bedroom", "Bedrooms"),
    ]
    contains_2_bed = False
    for field in bed_fields:
        if field and re.search(r"2\s*(?:bed|bedroom|br)", normalize_field(field), re.IGNORECASE):
            contains_2_bed = True
    if not contains_2_bed:
        issues.append("Property is not clearly a 2-bedroom unit.")

    bath_field = normalize_field(get_field_value(row, "Baths", "Bathroom", "Bathrooms") or "")
    if bath_field and re.search(r"(1|2)\s*(?:bath|bathroom|ba)", bath_field, re.IGNORECASE):
        if not re.search(r"2\s*(?:bath|bathroom|ba)", bath_field, re.IGNORECASE) and not re.search(r"1\s*(?:bath|bathroom|ba)", bath_field, re.IGNORECASE):
            issues.append("Bathroom count is not within the preferred 1-2 range.")

    if price_values:
        if max(price_values) > MAX_PRICE:
            issues.append(f"Price exceeds budget cap of ${MAX_PRICE}.")
    else:
        issues.append("Missing or invalid rent amount.")

    commute = parse_commute_minutes(
        get_field_value(row, "Commute", "Drive Time", "Commute Time") or ""
    )
    if commute is None:
        issues.append("Missing commute time to 08807.")
    elif commute > COMMUTE_LIMIT:
        issues.append(f"Commute exceeds {COMMUTE_LIMIT} minutes.")

    amenities_text = normalize_field(get_field_value(row, "Amenities", "Key Amenities") or "").lower()
    if "in-unit laundry" not in amenities_text and "washer" not in amenities_text and "dryer" not in amenities_text:
        issues.append("Missing in-unit laundry amenity.")
    if "disposal" not in amenities_text and "garbage disposal" not in amenities_text:
        issues.append("Missing garbage disposal amenity preference.")

    return {"passes": not issues, "issues": issues, "row": row}


def repair_workbook(path=EXCEL_PATH, output_path=None):
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=False)
    changed = False
    for sheet in workbook.worksheets:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if not first_row:
            continue
        keys = [str(cell or "").strip() for cell in first_row]
        for row in sheet.iter_rows(min_row=2, values_only=False):
            values = {}
            for idx in range(min(len(keys), len(row))):
                cell = row[idx]
                if cell._value is None:
                    continue
                if cell.coordinate not in sheet.merged_cells.ranges and hasattr(cell, "value"):
                    values[keys[idx]] = cell.value
            if not values or not any(value not in (None, "") for value in values.values()):
                continue
            repaired = repair_row(values)
            for idx, key in enumerate(keys[: len(row)]):
                if idx >= len(row):
                    continue
                cell = row[idx]
                if cell.coordinate in sheet.merged_cells.ranges:
                    continue
                new_value = repaired.get(key)
                if new_value is not None and cell.value != new_value:
                    cell.value = new_value
                    changed = True
    if output_path:
        workbook.save(output_path)
    elif changed:
        workbook.save(path)
    return {"changed": changed, "path": output_path or path}


def tavily_search(query):
    if not TAVILY_API_KEY:
        return []
    try:
        response = requests.post(
            "https://api.tavily.com/search",
            json={
                "api_key": TAVILY_API_KEY,
                "query": query,
                "max_results": 8,
                "search_depth": "basic",
                "include_answer": True,
                "include_raw_content": True,
            },
            timeout=45,
        )
        response.raise_for_status()
        payload = response.json()
        return payload.get("results", [])
    except Exception:
        return []


def openai_deep_dive(property_name, address, search_results):
    if not OPENAI_API_KEY:
        return {}
    candidate_text = "\n".join(
        f"- {item.get('title', '')}: {item.get('url', '')} | {str(item.get('content') or item.get('raw_content') or '')[:800]}"
        for item in search_results[:5]
    )
    try:
        response = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "model": "gpt-4o-mini",
                "messages": [
                    {"role": "system", "content": "You are validating apartment listings. Return valid JSON only."},
                    {"role": "user", "content": (
                        f"Validate this property against the user's rental profile: 2BR, 2BA preferred, max ${MAX_PRICE}, commute <= {COMMUTE_LIMIT} min, in-unit laundry, garbage disposal, managed community, official website, Bridgewater/Somerset NJ.\n"
                        f"Property name: {property_name}\nAddress hint: {address}\n"
                        "Return JSON only with keys: community_name, address, type, price, beds, baths, amenities, commute, phone, website, notes.\n"
                        f"Search results:\n{candidate_text}"
                    )},
                ],
                "temperature": 0.1,
            },
            timeout=60,
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"]
        if content.startswith("```"):
            content = re.sub(r"^```(?:json)?\s*|\s*```$", "", content.strip(), flags=re.IGNORECASE)
        parsed = json.loads(content)
        if isinstance(parsed, dict):
            return parsed
    except Exception:
        return {}
    return {}


def find_live_correction(row):
    property_name = normalize_field(
        get_field_value(row, "Community Name", "Property", "Name") or get_field_value(row, "Source") or ""
    )
    address = normalize_field(get_field_value(row, "Address", "Location") or "")
    if not property_name:
        return {}
    query = f'{property_name} {address} Bridgewater NJ 08807 official managed apartment website 2 bedroom in-unit laundry'
    results = tavily_search(query)
    if not results:
        return {}
    deep_dive = openai_deep_dive(property_name, address, results)
    if not deep_dive:
        return {}
    corrected = {}
    for key, value in {
        "Community Name": deep_dive.get("community_name") or deep_dive.get("property_name"),
        "Address": deep_dive.get("address"),
        "Property Type": deep_dive.get("type"),
        "Price": deep_dive.get("price"),
        "Beds": deep_dive.get("beds"),
        "Baths": deep_dive.get("baths"),
        "Amenities": deep_dive.get("amenities"),
        "Commute": deep_dive.get("commute"),
        "Leasing Office Phone": deep_dive.get("phone"),
        "Official Website": deep_dive.get("website"),
        "Notes": deep_dive.get("notes"),
    }.items():
        if value not in (None, ""):
            corrected[key] = value
    return corrected


def highlight_corrected_cells(sheet, row_index, corrected_names):
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    for header_name in corrected_names:
        lookup = {normalize_header(cell.value): idx for idx, cell in enumerate(sheet[1], start=1) if cell.value is not None}
        column_index = lookup.get(normalize_header(header_name))
        if column_index:
            target_cell = sheet.cell(row=row_index, column=column_index)
            target_cell.fill = green_fill


def ensure_validation_status_column(sheet):
    headers = [cell.value for cell in sheet[1]]
    normalized_headers = [normalize_header(header) for header in headers]
    if "validation status" not in normalized_headers:
        sheet.cell(row=1, column=sheet.max_column + 1, value="Validation Status")
        return sheet.max_column
    return normalized_headers.index("validation status") + 1


def validate_and_correct_workbook(path=EXCEL_PATH):
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=False)
    changed = False
    for sheet in workbook.worksheets:
        if sheet.max_row < 2:
            continue
        status_col = ensure_validation_status_column(sheet)
        headers = [cell.value for cell in sheet[1]]
        for row_index in range(2, sheet.max_row + 1):
            record = {}
            for col_index, header in enumerate(headers, start=1):
                value = sheet.cell(row=row_index, column=col_index).value
                record[str(header)] = value
            if not any(value not in (None, "") for value in record.values()):
                continue
            result = validate_row(record)
            if result["passes"]:
                sheet.cell(row=row_index, column=status_col, value="Pass")
                continue

            corrected = find_live_correction(record)
            if corrected:
                for key, value in corrected.items():
                    for col_index, header in enumerate(headers, start=1):
                        if normalize_header(header) == normalize_header(key):
                            if sheet.cell(row=row_index, column=col_index).value != value:
                                sheet.cell(row=row_index, column=col_index, value=value)
                                changed = True
                            break
                sheet.cell(row=row_index, column=status_col, value="Corrected")
                highlight_corrected_cells(sheet, row_index, corrected.keys())
            else:
                sheet.cell(row=row_index, column=status_col, value="Needs review")
        workbook.save(path)
    return {"path": path, "changed": changed}


def validate_workbook(path=EXCEL_PATH):
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    all_issues = []
    rows = []

    for sheet in workbook.worksheets:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if not first_row:
            continue
        keys = [str(cell or "").strip() for cell in first_row]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = {keys[idx]: value for idx, value in enumerate(row[: len(keys)])}
            if not any(value not in (None, "") for value in record.values()):
                continue
            result = validate_row(record)
            rows.append({"sheet": sheet.title, "record": record, "result": result})
            if not result["passes"]:
                all_issues.append({"sheet": sheet.title, "issues": result["issues"], "record": record})

    return {"pass": not all_issues, "issues": all_issues, "rows": rows}


def validate_csv(path):
    results = []
    with open(path, newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            result = validate_row(row)
            results.append(result)
    return {"pass": not any(not item["passes"] for item in results), "results": results}


if __name__ == "__main__":
    path = os.getenv("APARTMENT_EXCEL_PATH", EXCEL_PATH)
    fix_result = validate_and_correct_workbook(path)
    result = validate_workbook(path)
    print(f"Workbook updated: {fix_result['changed']}")
    print(f"Workbook validation passed: {result['pass']}")
    if result["issues"]:
        for issue in result["issues"][:10]:
            print(f"Sheet: {issue['sheet']}")
            print(f"Issues: {issue['issues']}")
            print(issue["record"])
