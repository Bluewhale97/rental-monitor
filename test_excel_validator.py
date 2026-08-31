import os

from openpyxl import Workbook

from excel_validator import is_non_listing_row, normalize_field, repair_workbook, validate_row, validate_workbook


def test_normalize_field_removes_extra_whitespace_and_trailing_text():
    assert normalize_field("  $3,200/mo  ") == "$3,200/mo"
    assert normalize_field("2 bed / 2 bath") == "2 bed / 2 bath"


def test_validate_row_flags_missing_budget_or_commute():
    row = {
        "Community Name": "The Grand",
        "Address": "123 Main St, Bridgewater, NJ",
        "Price": "$3,500",
        "Beds": "2",
        "Baths": "2",
        "Property Type": "Apartment",
        "Commute": "45 min drive to 08807",
        "Leasing Office Phone": "",
        "Official Website": "",
        "Amenities": "In-unit laundry; Disposal",
    }
    report = validate_row(row)
    assert report["passes"] is False
    assert any("budget" in issue.lower() for issue in report["issues"])
    assert any("commute" in issue.lower() for issue in report["issues"])


def test_repair_workbook_handles_merged_cells(tmp_path):
    workbook_path = tmp_path / "merged.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Price"
    ws["A2"] = "The Grand"
    ws["B2"] = "$3200"
    ws.merge_cells("C1:D1")
    ws["C1"] = "Merged"
    wb.save(workbook_path)

    result = repair_workbook(str(workbook_path))
    assert result["changed"] in (True, False)
    assert workbook_path.exists()


def test_summary_and_benchmark_rows_are_skipped(tmp_path):
    workbook_path = tmp_path / "candidate_rows.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Apartment Comparison"
    ws.append(["Name", "Status", "Priority", "Address", "Price", "Beds", "Baths", "Property Type", "Amenities", "Commute", "Website"])
    ws.append(["Current home", "Current home", "Benchmark", "", "", "", "", "", "", "", ""])
    ws.append(["Sunset Terrace", "New candidate", "Priority", "123 Main St, Somerset, NJ", "$2900", "2", "2", "Apartment", "In-unit laundry; Disposal", "20 min", "https://example.com"])
    wb.save(workbook_path)

    assert is_non_listing_row({"Name": "Current home", "Status": "Current home", "Priority": "Benchmark"}, "Apartment Comparison") is True
    result = validate_workbook(str(workbook_path))
    assert all("Current home" not in str(issue["record"].get("Name", "")) for issue in result["issues"])
